#!/usr/bin/env python3
"""
Export all scanned extensions from the database to an Excel file for QA and
competitive analysis. Includes scores, weightage, payload fields, and the
scoring model. Run from project root: uv run python scripts/export_qa_scoring_excel.py
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

# Project root and .env
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if (PROJECT_ROOT / ".env").exists():
    import dotenv
    dotenv.load_dotenv(PROJECT_ROOT / ".env")

# Add src so we can import extension_shield
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from extension_shield.api.database import db
from extension_shield.scoring.weights import (
    LAYER_WEIGHTS,
    SECURITY_WEIGHTS_V1,
    PRIVACY_WEIGHTS_V1,
    GOVERNANCE_WEIGHTS_V1,
)


def _flatten_scoring_v2(scoring_v2: dict | None) -> dict:
    """Flatten scoring_v2 into one-row dict for Excel."""
    out = {}
    if not scoring_v2 or not isinstance(scoring_v2, dict):
        return out
    out["overall_score"] = scoring_v2.get("overall_score")
    out["security_score"] = scoring_v2.get("security_score")
    out["privacy_score"] = scoring_v2.get("privacy_score")
    out["governance_score"] = scoring_v2.get("governance_score")
    out["base_overall"] = scoring_v2.get("base_overall")
    out["gate_penalty"] = scoring_v2.get("gate_penalty")
    gr = scoring_v2.get("gate_reasons")
    out["gate_reasons"] = " | ".join(gr) if isinstance(gr, list) else (gr or "")
    out["coverage_cap_applied"] = scoring_v2.get("coverage_cap_applied")
    out["coverage_cap_reason"] = scoring_v2.get("coverage_cap_reason") or ""
    out["risk_level"] = scoring_v2.get("risk_level")
    out["decision"] = scoring_v2.get("decision")
    gates = scoring_v2.get("hard_gates_triggered") or []
    out["hard_gates_triggered"] = ", ".join(gates) if isinstance(gates, list) else str(gates)
    out["weights_version"] = scoring_v2.get("weights_version")

    for layer_key, layer_label in [
        ("security_layer", "sec"),
        ("privacy_layer", "priv"),
        ("governance_layer", "gov"),
    ]:
        layer = scoring_v2.get(layer_key)
        if not isinstance(layer, dict):
            continue
        out[f"{layer_label}_layer_score"] = layer.get("score")
        factors = layer.get("factors") or []
        for f in factors:
            if not isinstance(f, dict):
                continue
            name = f.get("name") or "unknown"
            safe = name.replace(" ", "_")
            out[f"{layer_label}_{safe}_score"] = f.get("severity")  # 0-1 severity
            out[f"{layer_label}_{safe}_weight"] = f.get("weight")
            out[f"{layer_label}_{safe}_confidence"] = f.get("confidence")
    return out


def _our_verdict_bucket(decision: str, risk_level: str) -> str:
    """Map our decision/risk_level to Crxplorer-style bucket: Safe, Moderate, High Risk, Critical."""
    if not decision:
        return ""
    d = (decision or "").upper()
    r = (risk_level or "").upper()
    if d == "BLOCK":
        return "Critical" if r == "HIGH" else "High Risk"
    if d == "NEEDS_REVIEW":
        return "Moderate"
    return "Safe"


def _verdict_alignment(our_bucket: str, crx_verdict: str) -> str:
    """Compare our bucket to Crxplorer verdict: Aligned, We stricter, We looser, No data."""
    if not crx_verdict or not our_bucket:
        return "No data"
    crx = (crx_verdict or "").strip()
    if our_bucket == crx:
        return "Aligned"
    # Order: Safe < Moderate < High Risk < Critical
    order = {"Safe": 0, "Moderate": 1, "High Risk": 2, "Critical": 3}
    o_our = order.get(our_bucket, -1)
    o_crx = order.get(crx, -1)
    if o_our > o_crx:
        return "We stricter"
    if o_our < o_crx:
        return "We looser"
    return "Aligned"


def build_scans_rows(scans: list[dict]) -> list[dict]:
    """Build flat rows for the Scans sheet. Excludes high/medium/low_risk_count."""
    rows = []
    for s in scans:
        base = {
            "extension_name": s.get("extension_name") or "",
            "extension_url": s.get("url") or "",  # Chrome Web Store URL
            "extension_id": s.get("extension_id") or "",
            "timestamp": s.get("timestamp") or "",
            "security_score_legacy": s.get("security_score"),  # DB column
            "risk_level_legacy": s.get("risk_level"),
            "total_findings": s.get("total_findings"),
            "Crxplorer_score": "",   # e.g. 83
            "Crxplorer_verdict": "", # e.g. Safe, Moderate, High, Critical
            "Crxplorer_notes": "",
            "Extensionauditor_result": "",  # e.g. failed, Safe, Critical
            "Extensionauditor_notes": "",
            "Actual_issues": "",
            "Notes": "",
        }
        flat = _flatten_scoring_v2(s.get("scoring_v2"))
        base.update(flat)
        # Fallback: legacy rows without scoring_v2 may have security_score only
        if base.get("overall_score") is None and base.get("security_score_legacy") is not None:
            base["overall_score"] = base["security_score_legacy"]
        # Pre-fill competitor data where we have it (e.g. from user-provided QA input)
        _apply_competitor_prefill(base, s)

        # Comparison vs Crxplorer (for YC / QA)
        crx_score_raw = base.get("Crxplorer_score")
        try:
            crx_num = int(crx_score_raw) if crx_score_raw not in (None, "") else None
        except (TypeError, ValueError):
            crx_num = None
        our_score = base.get("overall_score")
        if our_score is not None:
            try:
                our_score = int(our_score)
            except (TypeError, ValueError):
                our_score = None
        base["Crxplorer_score_num"] = crx_num if crx_num is not None else ""
        if our_score is not None and crx_num is not None:
            base["score_diff"] = our_score - crx_num  # positive = we score higher
        else:
            base["score_diff"] = ""
        our_bucket = _our_verdict_bucket(
            base.get("decision") or "",
            base.get("risk_level") or "",
        )
        base["verdict_alignment"] = _verdict_alignment(our_bucket, base.get("Crxplorer_verdict") or "")
        # Short divergence note for slides
        if base.get("score_diff") != "" and base.get("verdict_alignment") != "No data":
            diff = base["score_diff"]
            sign = "+" if diff > 0 else ""
            base["divergence_notes"] = f"Our score {sign}{diff} vs Crxplorer; {base['verdict_alignment']}"
        else:
            base["divergence_notes"] = base.get("verdict_alignment") or ""

        rows.append(base)
    return rows


# Optional: pre-fill Crxplorer / Extensionauditor for known extensions (add more entries as you get data)
COMPETITOR_PREFILL = {
    "bmihdnoolhjbhjcoiamedcghafihcooi": {
        "Crxplorer_score": 83,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (AIS Visa Auto Rescheduler). Permissions 90% Low, Content Scripts 80% Low, WAR 90% Low, CSP 70% Medium, Externally Connectable 90% Low. Data Collection & Browser Access Medium Risk (visa details, ais.usvisa-info.com, cookies, storage; developer backend). Critical permissions confined to ais.usvisa-info.com and developer services.",
        "Extensionauditor_result": "failed",
        "Extensionauditor_notes": "Security analysis failed",
    },
    "kejbdjndbnbjgmefkgdddjlbokphdefk": {
        "Crxplorer_score": 92,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (Google). Permissions 90% Low, Content Scripts 95% Low, WAR 100% Low, CSP 70% Medium, Externally Connectable 95% Low. Data Collection & Browser Access Low Risk.",
        "Extensionauditor_result": "failed",
        "Extensionauditor_notes": "Security analysis failed",
    },
    "chphlpgkkbolifaimnlloiipkdnihall": {
        "Crxplorer_score": 85,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. Permissions 90% Low, Content Scripts 100% Low, WAR 100% Low, CSP 75% Low, Externally Connectable 100% Low. No content scripts/WAR/external connectivity. Data Collection & Browser Access Low Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "beepaenfejnphdgnkmccjcfiieihhogl": {
        "Crxplorer_score": 81,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. Scoped to usvisascheduling.com, atlasauth.b2clogin.com, checkvisaslots.com. Permissions 90% Low, Content Scripts 95% Low, WAR 100% Low, CSP 80% Low, Externally Connectable 90% Low. Data Collection Medium Risk, Browser Access Low Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "kpdimldbpabboiekkcbcffgmamoenkom": {
        "Crxplorer_score": 99,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. No risky permissions, content scripts, WAR, or external connectivity. Permissions 100% Low, Content Scripts 100% Low, WAR 100% Low, CSP 95% Low, Externally Connectable 100% Low. Data Collection & Browser Access Low Risk; self-contained.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "ebohipfojjgamhnfgdohlfkfagianjim": {
        "Crxplorer_score": 95,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. Minimal manifest: no special permissions, no content scripts, no WAR, no external connectivity. All categories 100% Low Risk. Data Collection & Browser Access Low Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "ndnaehgpjlnokgebbaldlmgkapkpjkkb": {
        "Crxplorer_score": 82,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (Mailtrack/Mailsuite). Permissions 95% Low, Content Scripts 90% Low, CSP 70% Medium, Externally Connectable 100% Low, WAR 95% Low. Data Collection High Risk (email tracking, Gmail content); Browser Access Medium Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "mdanidgdpmkimeiiojknlnekblgmpdll": {
        "Crxplorer_score": 75,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (Boomerang). Permissions 40% High Risk, Content Scripts 95% Low, WAR 90% Low, CSP 30% High Risk, Externally Connectable 95% Low. Data Collection & Browser Access Medium Risk; 'management' permission; scoped to mail.google.com.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "gpoaahhbceojgenmhabidkbglkifibah": {
        "Crxplorer_score": 79,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. Limited scope (etherscan.io/tx/*). Permissions 70% Medium, Content Scripts 80% Low, WAR 100% Low, CSP 70% Medium, Externally Connectable 100% Low. Data Collection Medium Risk (Etherscan tx page data); Browser Access Low Risk. No WAR; not externally connectable.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "adbacgifemdbhdkfppmeilbgppmhaobf": {
        "Crxplorer_score": 81,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (RoPro). Permissions 90% Low, Content Scripts 85% Low, WAR 90% Low, CSP 60% Medium, Externally Connectable 100% Low. Data Collection & Browser Access Medium Risk (roblox.com, ropro.io, storage). externally_connectable empty.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "glelnofehkiolecgdagikdfkfhoafogd": {
        "Crxplorer_score": 20,
        "Crxplorer_verdict": "Critical",
        "Crxplorer_notes": "Security Alert. WAR 10% Critical (broad web_accessible_resources). Permissions 95% Low, Content Scripts 90% Low, CSP 60% Medium, Externally Connectable 80% Low. Data Collection & Browser Access High Risk (Gmail, Outlook, Salesforce). Do not install recommended.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "fdpohaocaechififmbbbbbknoalclacl": {
        "Crxplorer_score": 59,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended. Permissions 90% Low, Content Scripts 95% Low, WAR 70% Medium, CSP 10% Critical (Externally Connectable), Externally Connectable 95% Low. Data Collection Low Risk; Browser Access Medium Risk. activeTab/storage/unlimitedStorage; no auto content scripts; potentially gofullpage.com.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "kojibkalenabblnhoihknojdfapbbmig": {
        "Crxplorer_score": 63,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended (Snipo). Permissions 90% Low, Content Scripts 80% Low, WAR 95% Low, CSP 60% Medium, Externally Connectable 100% Low. Data Collection & Browser Access High Risk (tabs, *://*/*, clipboardRead, storage; Notion/YouTube/Skillshare/Vimeo). WAR scoped to first-party; no critical manifest issues.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "bdlffpdkioakgjjpmgpboogfiaegbpgp": {
        "Crxplorer_score": 98,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. Minimal permissions (storage, notifications, alarms). No content scripts, no WAR, no externally_connectable. Permissions 95% Low, Content Scripts 100% Low, WAR 100% Low, CSP 95% Low, Externally Connectable 100% Low. Data Collection & Browser Access Low Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "bamdkjfjhhnjcgcjmmjdnncpglihepoi": {
        "Crxplorer_score": 98,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. No permissions, content scripts, WAR, or external connectivity. Permissions/Content Scripts/WAR 100% Low, CSP 95% Low, Externally Connectable 100% Low. Data Collection & Browser Access Low Risk (none). Functionally inert per manifest.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "mjcklhnhfiepmofggcoegkmkokbljmjd": {
        "Crxplorer_score": 54,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended. No CSP (0% Critical); localhost in WAR and externally_connectable; *:// for domains (HTTP risk). Permissions 85% Low, Content Scripts 90% Low, WAR 50% Medium, Externally Connectable 40% High. Data Collection & Browser Access Medium (topSites, Google OAuth, battletabs.*).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "abbpaepbpakcpipajigmlpnhlnbennna": {
        "Crxplorer_score": 86,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. Permissions 90% Low, Content Scripts 95% Low, WAR 90% Low, CSP 60% Medium, Externally Connectable 100% Low. Data Collection & Browser Access Medium Risk (YouTube URLs/content, storage, declarativeNetRequest; modify YouTube/YouTube Music). WAR scoped to YouTube domains.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "jiihfaimpblhjckejcegajknncfamimb": {
        "Crxplorer_score": 85,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified. No special permissions, minimal manifest. No content scripts, WAR, or external communication. Permissions/Content Scripts/WAR 100% Low, CSP 80% Low, Externally Connectable 100% Low. Data Collection & Browser Access Low Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "hjngolefdpdnooamgdldlkjgmdcmcjnc": {
        "Crxplorer_score": 73,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended (Texthelp). Negative user feedback (spam spyware, can't remove). Permissions 85% Low, Content Scripts 90% Low, WAR 90% Low, CSP 35% High Risk (Externally Connectable), Externally Connectable 90% Low. Data Collection & Browser Access Medium (Docs/Office/Texthelp/Coursera, OAuth, GCM, equatio.texthelp.com). unsafe-eval CSP.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "ecnphlgnajanjnkcmbpancdjoidceilk": {
        "Crxplorer_score": 98,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (Kami, ed-tech). Permissions 90% Low, Content Scripts 95% Low, WAR 90% Low, CSP 65% Medium, Externally Connectable 95% Low. Data Collection & Browser Access Low Risk. WAR and external connectivity restricted to own domains and educational platforms.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "hoombieeljmmljlkjmnheibnpciblicm": {
        "Crxplorer_score": 67,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended. Broad *://*.amazon.*/* content script. Permissions 90% Low, Content Scripts 30% High Risk, WAR 100% Low, CSP 30% High Risk, Externally Connectable 100% Low. Data Collection & Browser Access High Risk (Netflix, YouTube, Amazon; languagereactor.com).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "eiaeiblijfjekdanodkjadfinkhbfgcd": {
        "Crxplorer_score": 75,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (NordPass, Nord Security). Permissions 95% Low, Content Scripts 80% Low, WAR 40% High Risk, CSP 0% Critical, Externally Connectable 100% Low. Data Collection & Browser Access High Risk (passwords, form data, URLs; inject on any page). Missing CSP; broad content scripts necessary for password manager.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "likadllkkidlligfcdhfnnbkjigdkmci": {
        "Crxplorer_score": 53,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended (QRExt). Unclear purpose; tabs permission invasive; no CSP. Permissions 50% Medium, Content Scripts 90% Low, CSP 30% High Risk, Externally Connectable 100% Low, WAR 100% Low. Data Collection & Browser Access High Risk (all tab URLs/favicons, activeTab, storage/unlimitedStorage).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "aapbdbdomjkkjkaonfhkkikfgjllcleb": {
        "Crxplorer_score": 73,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended (Google). Minimal permissions (activeTab, contextMenus, storage); no content_scripts, no WAR, no externally_connectable. Permissions 95% Low, Content Scripts/WAR/CSP/Externally Connectable 100% or 95% Low. Data Collection & Browser Access Low Risk (active tab on invoke, local storage).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "bmnlcjabgnpnenekpadlanbbkooimhnj": {
        "Crxplorer_score": 60,
        "Crxplorer_verdict": "Moderate",
        "Crxplorer_notes": "Review Recommended (PayPal Honey). Negative reviews (scam, stealing information). WAR broad (http(s)://*/*) 30% High Risk; no custom CSP. Permissions 90% Low, Content Scripts 90% Low, CSP 60% Medium, Externally Connectable 90% Low. Data Collection & Browser Access High Risk (any site, cookies, webRequest, storage).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "pgbdljpkijehgoacbjpolaomhkoffhnl": {
        "Crxplorer_score": 87,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (MailTracker). Gmail-specific; no WAR, no external connections. Permissions 90% Low, Content Scripts 95% Low, WAR 100% Low, CSP 70% Medium, Externally Connectable 100% Low. Data Collection & Browser Access High Risk (Gmail content, interactions, network; inject/modify Gmail, webRequest, storage).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "okkpmekocmfambbdiieimopoccpdfbef": {
        "Crxplorer_score": 95,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (Hipcamp). No special permissions, content scripts, or external interaction. All categories 100% Low Risk. Data Collection & Browser Access Low Risk (none declared). Minimal manifest; new tab camping inspiration.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "nkabooldphfdjcbhcodblkfmigmpchhi": {
        "Crxplorer_score": 32,
        "Crxplorer_verdict": "High Risk",
        "Crxplorer_notes": "Caution Advised (Vitaminiser). Broad permissions without justified purpose: content scripts on all HTTP/HTTPS, WAR for all sites, cookies. Permissions 45% High Risk, Content Scripts 15% Critical, WAR 5% Critical, CSP 70% Medium. Data Collection & Browser Access High Risk. Do not install recommended.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "abikfbojmghmfjdjlbagiamkinbmbaic": {
        "Crxplorer_score": 28,
        "Crxplorer_verdict": "High Risk",
        "Crxplorer_notes": "Caution Advised. tabCapture (picture/video of tab) excessive for audio equalizer; WAR *://*/* 10% Critical; no CSP; externally_connectable to addonup.com. Permissions 15% Critical, Content Scripts 70% Medium, Externally Connectable 45% High Risk. Data Collection & Browser Access High Risk.",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "ljphffonbjcankbnpmpmlpjnnopdjkkf": {
        "Crxplorer_score": 82,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (WABlasters). WhatsApp Web–focused; no external connections; WAR scoped to WhatsApp. Permissions 95% Low, Content Scripts 90% Low, WAR 90% Low, CSP 60% Medium, Externally Connectable 95% Low. Data Collection & Browser Access Medium Risk (web.whatsapp.com, storage).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
    "oligonmocnihangdjlloenpndnniikol": {
        "Crxplorer_score": 78,
        "Crxplorer_verdict": "Safe",
        "Crxplorer_notes": "Extension Verified (Edpuzzle). Edpuzzle/YouTube educational; tabCapture, desktopCapture; no WAR. Permissions 85% Low, Content Scripts 95% Low, WAR 100% Low, CSP 80% Low, Externally Connectable 95% Low. Data Collection & Browser Access Medium Risk (tabs, tabCapture, desktopCapture, edpuzzle.com/youtube.com, storage).",
        "Extensionauditor_result": "",
        "Extensionauditor_notes": "",
    },
}


def _apply_competitor_prefill(base: dict, scan: dict) -> None:
    """If this scan matches a known extension in COMPETITOR_PREFILL, fill in competitor columns."""
    ext_id = (scan.get("extension_id") or "").strip()
    url = (scan.get("url") or "").strip()
    name = (scan.get("extension_name") or "").strip()
    key = None
    if ext_id and ext_id in COMPETITOR_PREFILL:
        key = ext_id
    elif url and "bmihdnoolhjbhjcoiamedcghafihcooi" in url and "bmihdnoolhjbhjcoiamedcghafihcooi" in COMPETITOR_PREFILL:
        key = "bmihdnoolhjbhjcoiamedcghafihcooi"
    if not key:
        return
    prefill = COMPETITOR_PREFILL.get(key, {})
    for k, v in prefill.items():
        if k in base and base.get(k) == "":
            base[k] = v


def main() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("Install openpyxl: uv sync --group dev  (or pip install openpyxl)")
        sys.exit(1)

    out_path = PROJECT_ROOT / "qa_scoring_export.xlsx"
    print(f"Fetching all completed scans from database...")
    scans = db.get_recent_scans(limit=10000)
    if not scans:
        print("No completed scans found in the database.")
        sys.exit(0)
    print(f"Found {len(scans)} scan(s). Building Excel...")

    rows = build_scans_rows(scans)
    if not rows:
        print("No rows to write.")
        sys.exit(0)

    wb = Workbook()
    ws_scans = wb.active
    ws_scans.title = "Scans"

    # Headers from first row keys (consistent order)
    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws_scans.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            val = row.get(h)
            if isinstance(val, (dict, list)):
                val = json.dumps(val)[:32767]  # Excel cell limit
            ws_scans.cell(row=row_idx, column=col_idx, value=val)

    # Sheet 2: Scoring model (weightage)
    ws_model = wb.create_sheet("Scoring model", 1)
    ws_model.cell(1, 1, "Layer (overall)").font = Font(bold=True)
    ws_model.cell(1, 2, "Weight").font = Font(bold=True)
    for i, (name, w) in enumerate(LAYER_WEIGHTS.items(), 2):
        ws_model.cell(i, 1, name)
        ws_model.cell(i, 2, w)
    row = 2 + len(LAYER_WEIGHTS) + 1
    ws_model.cell(row, 1, "Security layer factors").font = Font(bold=True)
    row += 1
    for name, w in SECURITY_WEIGHTS_V1.items():
        ws_model.cell(row, 1, name)
        ws_model.cell(row, 2, w)
        row += 1
    ws_model.cell(row, 1, "Privacy layer factors").font = Font(bold=True)
    row += 1
    for name, w in PRIVACY_WEIGHTS_V1.items():
        ws_model.cell(row, 1, name)
        ws_model.cell(row, 2, w)
        row += 1
    ws_model.cell(row, 1, "Governance layer factors").font = Font(bold=True)
    row += 1
    for name, w in GOVERNANCE_WEIGHTS_V1.items():
        ws_model.cell(row, 1, name)
        ws_model.cell(row, 2, w)
        row += 1
    row += 1
    ws_model.cell(row, 1, "Hard gate penalties (base penalty points, applied to layer score)").font = Font(bold=True)
    row += 1
    gate_penalty_rows = [
        ("Gate ID", "Layer", "Base penalty", "Notes"),
        ("CRITICAL_SAST", "security", 50, "BLOCK gate"),
        ("VT_MALWARE", "security", 45, "BLOCK gate"),
        ("TOS_VIOLATION", "governance", 60, "BLOCK gate"),
        ("PURPOSE_MISMATCH", "governance", 45, "WARN/BLOCK gate"),
        ("SENSITIVE_EXFIL", "privacy", 40, "WARN gate"),
    ]
    for i, r in enumerate(gate_penalty_rows):
        for c, val in enumerate(r, 1):
            ws_model.cell(row, c, val)
        if i == 0:
            for c in range(1, 5):
                ws_model.cell(row, c).font = Font(bold=True)
        row += 1
    row += 1
    ws_model.cell(row, 1, "Overall: base_overall = weighted layer sum (pre-penalty). Penalties applied per layer (max per layer); overall_after_gates = weighted sum(adjusted layers). final_overall = min(overall_after_gates, 80) if coverage_cap else overall_after_gates.")
    row += 1
    ws_model.cell(row, 1, "Coverage cap: when SAST missing (0 files scanned, no findings), overall is capped at 80 and decision at least NEEDS_REVIEW.")

    # Sheet 3: What we have vs competitors
    ws_diff = wb.create_sheet("What we have vs competitors", 2)
    ws_diff.cell(1, 1, "Capability").font = Font(bold=True)
    ws_diff.cell(1, 2, "ExtensionShield").font = Font(bold=True)
    ws_diff.cell(1, 3, "Notes (Crxplorer / Extensionauditor)").font = Font(bold=True)
    diff_rows = [
        ("Semgrep SAST", "Yes – custom rules, 47+", "Compare SAST coverage"),
        ("VirusTotal", "Yes – malware hash check", "Optional; we surface engine counts"),
        ("Obfuscation detection", "Yes – factor + weight", "Often missing elsewhere"),
        ("Manifest / CSP", "Yes – factor", "CSP missing often flagged by us"),
        ("ChromeStats / behavioral", "Yes – factor", "Behavioral intel"),
        ("Webstore / Maintenance", "Yes – factors", "Rating, freshness, privacy policy"),
        ("Privacy layer (Permissions, Combos, NetworkExfil, Capture)", "Yes – 4 factors", "Compare permission vs exfil focus"),
        ("Governance layer (ToS, Consistency, Disclosure)", "Yes – 3 factors", "Policy alignment"),
        ("Layer weights (Security 50%, Privacy 30%, Gov 20%)", "Yes – configurable", "Transparent in this sheet"),
        ("Hard gates (e.g. VT malware → BLOCK)", "Yes", "Binary block vs score-only"),
        ("Evidence / chain-of-custody", "Yes – governance_bundle", "For audits"),
    ]
    for i, (cap, us, notes) in enumerate(diff_rows, 2):
        ws_diff.cell(i, 1, cap)
        ws_diff.cell(i, 2, us)
        ws_diff.cell(i, 3, notes)

    # Sheet 4: YC Comparison – ExtensionShield vs Crxplorer (32 extensions with Crxplorer prefill)
    ws_yc = wb.create_sheet("YC Comparison", 3)
    row = 1
    ws_yc.cell(row, 1, "ExtensionShield vs Crxplorer – Summary (for YC / investor deck)").font = Font(bold=True, size=12)
    row += 2

    # Rows that have both our score and Crxplorer score
    comparable = [r for r in rows if r.get("Crxplorer_score_num") != "" and r.get("overall_score") is not None]
    n_comp = len(comparable)
    verdict_counts = {}
    for r in comparable:
        v = r.get("verdict_alignment") or "No data"
        verdict_counts[v] = verdict_counts.get(v, 0) + 1
    mean_diff = sum(r.get("score_diff", 0) or 0 for r in comparable) / n_comp if n_comp else 0
    stricter = verdict_counts.get("We stricter", 0)
    looser = verdict_counts.get("We looser", 0)
    aligned = verdict_counts.get("Aligned", 0)

    summary_data = [
        ("Metric", "Value"),
        ("Extensions with Crxplorer prefill (compared)", n_comp),
        ("Mean score difference (Our score − Crxplorer)", round(mean_diff, 1)),
        ("Verdict Aligned", aligned),
        ("We stricter (we block/warn, they safe/moderate)", stricter),
        ("We looser (we allow, they high/critical)", looser),
    ]
    for i, (label, val) in enumerate(summary_data, row):
        ws_yc.cell(i, 1, label)
        ws_yc.cell(i, 2, val)
        if i == row:
            ws_yc.cell(i, 1, label).font = Font(bold=True)
            ws_yc.cell(i, 2, val).font = Font(bold=True)
    row += len(summary_data) + 2

    ws_yc.cell(row, 1, "Findings – Where ExtensionShield is stronger").font = Font(bold=True)
    row += 1
    findings_us = [
        "Semgrep SAST (47+ rules) + VirusTotal malware gate – we block on detection.",
        "Obfuscation detection as a weighted factor; Crxplorer does not surface it.",
        "Hard gates (VT malware → BLOCK); Crxplorer is score-only, no binary block.",
        "Privacy layer: permission combos, network exfil, capture (mic/camera) as factors.",
        "Governance: ToS consistency, disclosure; evidence bundle for audits.",
        "Transparent layer weights (Security 50%, Privacy 30%, Gov 20%) in Scoring model sheet.",
    ]
    for line in findings_us:
        ws_yc.cell(row, 1, f"• {line}")
        row += 1
    row += 1

    ws_yc.cell(row, 1, "Where Crxplorer adds value (we could adopt)").font = Font(bold=True)
    row += 1
    findings_them = [
        "5-category breakdown (Permissions, Content Scripts, WAR, CSP, Externally Connectable) with % risk – good for UX.",
        "Impact analysis (Data Collection / Browser Access risk level) – we have factors but could surface similarly.",
        "User-review sentiment mentioned in key findings – we have webstore_analysis but could highlight more.",
    ]
    for line in findings_them:
        ws_yc.cell(row, 1, f"• {line}")
        row += 1
    row += 1

    ws_yc.cell(row, 1, "Pending in our scoring (improvements)").font = Font(bold=True)
    row += 1
    pending = [
        "Explicit CSP strength score (we have factor; could mirror their 0–100% per category).",
        "Externally connectable allowlist as dedicated factor if not covered.",
        "Optional: verdict buckets (Safe / Moderate / High / Critical) for direct Crxplorer comparison.",
    ]
    for line in pending:
        ws_yc.cell(row, 1, f"• {line}")
        row += 1
    row += 2

    # Chart data: Extension name | Our score | Crxplorer score (comparable only)
    ws_yc.cell(row, 1, "Chart data – Score comparison").font = Font(bold=True)
    row += 1
    ws_yc.cell(row, 1, "Extension name")
    ws_yc.cell(row, 2, "Our score")
    ws_yc.cell(row, 3, "Crxplorer score")
    for c in range(1, 4):
        ws_yc.cell(row, c).font = Font(bold=True)
    row += 1
    chart_start_row = row
    for r in comparable:
        name = (r.get("extension_name") or "Unknown")[:40]
        ws_yc.cell(row, 1, name)
        ws_yc.cell(row, 2, r.get("overall_score"))
        ws_yc.cell(row, 3, r.get("Crxplorer_score_num"))
        row += 1
    chart_end_row = row - 1

    # Add bar chart (Our score vs Crxplorer score by extension)
    if n_comp > 0 and chart_end_row >= chart_start_row:
        try:
            from openpyxl.chart import BarChart, Reference
            chart = BarChart(type="col", title="ExtensionShield vs Crxplorer score (same extensions)")
            chart.y_axis.title = "Score (0–100)"
            chart.x_axis.title = "Extension"
            data = Reference(ws_yc, min_col=2, min_row=chart_start_row - 1, max_row=chart_end_row, max_col=3)
            cats = Reference(ws_yc, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 18
            chart.height = 12
            ws_yc.add_chart(chart, f"E{chart_start_row}")
        except Exception as chart_err:
            print(f"Note: Could not add chart to YC Comparison sheet: {chart_err}")

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print("Sheets: Scans, Scoring model, What we have vs competitors, YC Comparison (summary + chart).")


if __name__ == "__main__":
    main()
