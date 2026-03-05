#!/usr/bin/env python3
"""
QA regression: compare a baseline list of extensions/scores to the current export.

Use this to ensure the same set of extensions is still scanned and (optionally)
that scores/decisions have not regressed. Baseline can be the verification CSV
or the Excel; current is the Excel (fresh export from DB).

Usage (from project root):
  uv run python scripts/qa_regression_scoring.py --baseline docs/qa_extensionshield/qa_scoring_verification_report.csv --current docs/qa_extensionshield/qa_scoring_export.xlsx
  uv run python scripts/qa_regression_scoring.py --baseline docs/qa_extensionshield/qa_scoring_export.xlsx --current docs/qa_extensionshield/qa_scoring_export.xlsx  # same file = self-check

Exit code: 0 if no regressions (or only allowed additions), 1 if baseline extensions missing or score/decision regressions.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))


def read_csv_baseline(csv_path: Path) -> list[dict]:
    """Read baseline from verification CSV. Keys: extension_id, our_overall, our_decision, our_name."""
    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            eid = (row.get("extension_id") or "").strip()
            if eid:
                rows.append({
                    "extension_id": eid,
                    "our_name": row.get("our_name") or "",
                    "our_overall": _num(row.get("our_overall")),
                    "our_decision": (row.get("our_decision") or "").strip(),
                })
    return rows


def read_excel_scans(excel_path: Path) -> list[dict]:
    """Read Scans sheet from QA export Excel."""
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: uv sync (or pip install openpyxl)")
        sys.exit(1)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "Scans" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Scans"]
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = list(headers_row) if headers_row else []
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        eid = (r.get("extension_id") or "").strip()
        if eid:
            rows.append({
                "extension_id": eid,
                "our_name": r.get("extension_name") or "",
                "our_overall": _num(r.get("overall_score")),
                "our_decision": (r.get("decision") or "").strip(),
            })
    wb.close()
    return rows


def _num(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def run_regression(
    baseline_path: Path,
    current_path: Path,
    baseline_is_csv: bool,
    current_is_csv: bool,
    strict_scores: bool,
) -> int:
    if baseline_is_csv:
        baseline = read_csv_baseline(baseline_path)
    else:
        baseline = read_excel_scans(baseline_path)
    if current_is_csv:
        current = read_csv_baseline(current_path)
    else:
        current = read_excel_scans(current_path)

    if not baseline:
        print(f"Baseline has no extensions: {baseline_path}")
        return 1
    if not current:
        print(f"Current has no extensions: {current_path}")
        return 1

    base_ids = {r["extension_id"] for r in baseline}
    cur_ids = {r["extension_id"] for r in current}
    cur_by_id = {r["extension_id"]: r for r in current}

    missing = base_ids - cur_ids
    added = cur_ids - base_ids
    common = base_ids & cur_ids

    exit_code = 0
    regressions = []
    print("=== QA Regression: baseline vs current ===\n")
    print(f"Baseline: {baseline_path} ({len(baseline)} extensions)")
    print(f"Current:  {current_path} ({len(current)} extensions)\n")

    if missing:
        exit_code = 1
        print(f"REGRESSION: {len(missing)} extension(s) from baseline missing in current:")
        for eid in sorted(missing):
            name = next((r["our_name"] for r in baseline if r["extension_id"] == eid), "")
            print(f"  - {eid}  {name[:50]}")
        print()
    else:
        print("OK: All baseline extensions present in current.\n")

    if added:
        print(f"ADDED: {len(added)} extension(s) in current not in baseline:")
        for eid in sorted(added):
            r = cur_by_id.get(eid, {})
            print(f"  + {eid}  {(r.get('our_name') or '')[:50]}")
        print()

    if strict_scores and common:
        print("Score/decision comparison (baseline vs current):")
        base_by_id = {r["extension_id"]: r for r in baseline}
        for eid in sorted(common):
            b, c = base_by_id[eid], cur_by_id[eid]
            bo, co = b.get("our_overall"), c.get("our_overall")
            bd, cd = b.get("our_decision") or "", c.get("our_decision") or ""
            if bo is not None and co is not None and bo != co:
                regressions.append((eid, f"overall {bo} -> {co}", b.get("our_name")))
            elif bd != cd:
                regressions.append((eid, f"decision {bd} -> {cd}", b.get("our_name")))
        if regressions:
            exit_code = 1
            for eid, msg, name in regressions:
                print(f"  CHANGE: {eid}  {(name or '')[:40]}  {msg}")
        else:
            print("  OK: No score/decision changes in common extensions.")
    elif strict_scores:
        print("(No common extensions to compare scores.)")

    print("\n=== Summary ===")
    print(f"  Baseline count: {len(baseline)}")
    print(f"  Current count:  {len(current)}")
    print(f"  Missing (in baseline, not in current): {len(missing)}")
    print(f"  Added (in current, not in baseline):   {len(added)}")
    if strict_scores and common:
        print(f"  Score/decision regressions: {len(regressions)}")

    return exit_code


def main() -> None:
    parser = argparse.ArgumentParser(description="Regression: compare baseline vs current QA export")
    parser.add_argument("--baseline", type=Path, required=True, help="Baseline file (CSV or Excel)")
    parser.add_argument("--current", type=Path, required=True, help="Current export (CSV or Excel)")
    parser.add_argument("--strict", action="store_true", help="Fail on score/decision changes in common extensions")
    args = parser.parse_args()

    def is_csv(p: Path) -> bool:
        return p.suffix.lower() == ".csv"

    if not args.baseline.exists():
        print(f"Baseline not found: {args.baseline}")
        sys.exit(1)
    if not args.current.exists():
        print(f"Current not found: {args.current}")
        sys.exit(1)

    code = run_regression(
        args.baseline,
        args.current,
        baseline_is_csv=is_csv(args.baseline),
        current_is_csv=is_csv(args.current),
        strict_scores=args.strict,
    )
    sys.exit(code)


if __name__ == "__main__":
    main()
