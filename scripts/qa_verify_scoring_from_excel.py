#!/usr/bin/env python3
"""
QA: Verify scoring results from qa_scoring_export.xlsx against Chrome Web Store.

Reads docs/qa_extensionshield/qa_scoring_export.xlsx (sheet "Scans"), fetches
each extension's store listing, and produces a report comparing our name/score
with the store's title, developer, rating, and user count.

Usage (from project root):
  uv run python scripts/qa_verify_scoring_from_excel.py
  uv run python scripts/qa_verify_scoring_from_excel.py --limit 5   # first 5 only
  uv run python scripts/qa_verify_scoring_from_excel.py --out report.md

Requires: openpyxl, requests, beautifulsoup4 (already in deps). Network access for CWS.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if (PROJECT_ROOT / ".env").exists():
    import dotenv
    dotenv.load_dotenv(PROJECT_ROOT / ".env")
sys.path.insert(0, str(PROJECT_ROOT / "src"))

# ---------------------------------------------------------------------------
# Read Excel
# ---------------------------------------------------------------------------

def read_qa_excel(excel_path: Path, limit: int | None = None) -> list[dict]:
    """Read Scans sheet from QA export Excel. Returns list of row dicts."""
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: uv sync (or pip install openpyxl)")
        sys.exit(1)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "Scans" not in wb.sheetnames:
        print(f"Sheet 'Scans' not found in {excel_path}. Sheets: {wb.sheetnames}")
        wb.close()
        return []
    ws = wb["Scans"]
    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = list(headers_row) if headers_row else []
    rows = []
    for i, row in enumerate(rows_iter):
        if limit is not None and i >= limit:
            break
        r = dict(zip(headers, row))
        if r.get("extension_id"):
            rows.append(r)
    wb.close()
    return rows


def _norm(s: str | None) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().lower().split())


def _name_match(our: str | None, store: str | None) -> str:
    if not our and not store:
        return "—"
    if not our or not store:
        return "missing"
    a, b = _norm(our), _norm(store)
    if a == b:
        return "match"
    if a in b or b in a:
        return "partial"
    return "mismatch"


# ---------------------------------------------------------------------------
# Fetch store metadata
# ---------------------------------------------------------------------------

def fetch_store_metadata(extension_id: str):
    """Fetch Chrome Web Store metadata for extension_id. Returns dict or None."""
    # CWS accepts /detail/_/id and redirects to the slugged URL
    store_url = f"https://chromewebstore.google.com/detail/_/{extension_id}"
    try:
        from extension_shield.core.extension_metadata import ExtensionMetadata
        fetcher = ExtensionMetadata(store_url)
        return fetcher.fetch_metadata()
    except Exception as e:
        return {"_error": str(e)}


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def run_verification(excel_path: Path, limit: int | None, out_path: Path | None) -> None:
    rows = read_qa_excel(excel_path, limit=limit)
    if not rows:
        print("No rows with extension_id found.")
        return
    print(f"Loaded {len(rows)} extension(s) from {excel_path}. Fetching store data...")
    results = []
    for i, r in enumerate(rows):
        ext_id = (r.get("extension_id") or "").strip()
        our_name = r.get("extension_name") or ""
        our_overall = r.get("overall_score")
        our_decision = r.get("decision") or ""
        our_risk = r.get("risk_level") or ""
        sec = r.get("security_score")
        priv = r.get("privacy_score")
        gov = r.get("governance_score")
        if not ext_id:
            continue
        meta = fetch_store_metadata(ext_id)
        fetch_error = ""
        if meta and meta.get("_error"):
            store_title = ""
            store_developer = ""
            store_rating = ""
            store_users = ""
            fetch_error = meta["_error"][:120]
            name_check = "error"
        else:
            store_title = (meta or {}).get("title") or ""
            store_developer = (meta or {}).get("developer_name") or ""
            store_rating = (meta or {}).get("rating")
            store_users = (meta or {}).get("user_count")
            if store_rating is not None:
                store_rating = f"{store_rating:.1f}"
            else:
                store_rating = ""
            if store_users is not None:
                store_users = f"{store_users:,}"
            else:
                store_users = ""
            name_check = _name_match(our_name, store_title)
        results.append({
            "extension_id": ext_id,
            "our_name": our_name,
            "store_title": store_title,
            "store_developer": store_developer,
            "name_check": name_check,
            "fetch_error": fetch_error,
            "our_overall": our_overall,
            "our_decision": our_decision,
            "our_risk": our_risk,
            "security_score": sec,
            "privacy_score": priv,
            "governance_score": gov,
            "store_rating": store_rating if not fetch_error else "",
            "store_users": store_users if not fetch_error else "",
        })
        print(f"  [{i+1}/{len(rows)}] {ext_id} -> {store_title or '(no title)'} | name_check={name_check}")

    # Write report
    if out_path and out_path.suffix.lower() == ".md":
        with open(out_path, "w") as f:
            f.write("# QA: Scoring verification vs Chrome Web Store\n\n")
            f.write(f"Source: `{excel_path}` ({len(results)} extensions)\n\n")
            f.write("| Extension ID | Our name | Store title | Name check | Our overall | Decision | Store rating | Store users |\n")
            f.write("|--------------|----------|-------------|------------|--------------|----------|--------------|-------------|\n")
            for r in results:
                our = (r["our_name"] or "").replace("|", ",")[:30]
                store = (r["store_title"] or "").replace("|", ",")[:30]
                f.write(f"| {r['extension_id']} | {our} | {store} | {r['name_check']} | {r['our_overall']} | {r['our_decision']} | {r['store_rating']} | {r['store_users']} |\n")
            f.write("\n## Summary\n\n")
            matches = sum(1 for x in results if x["name_check"] == "match")
            partial = sum(1 for x in results if x["name_check"] == "partial")
            mismatch = sum(1 for x in results if x["name_check"] == "mismatch")
            errors = sum(1 for x in results if x.get("fetch_error"))
            f.write(f"- Name match: {matches}, partial: {partial}, mismatch: {mismatch}\n")
            f.write(f"- Store fetch errors or missing title: {errors}\n")
        print(f"Wrote {out_path}")
    elif out_path and out_path.suffix.lower() == ".csv":
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=results[0].keys() if results else [])
            w.writeheader()
            w.writerows(results)
        print(f"Wrote {out_path}")
    else:
        # Default: print table to stdout
        print("\n" + "=" * 100)
        print(f"{'extension_id':<34} {'our_name':<28} {'store_title':<28} {'name':<8} {'overall':<6} {'decision':<12} {'store_rating':<6} {'store_users':<12}")
        print("=" * 100)
        for r in results:
            print(f"{r['extension_id']:<34} {(r['our_name'] or '')[:27]:<28} {(r['store_title'] or '')[:27]:<28} {r['name_check']:<8} {str(r['our_overall']):<6} {r['our_decision']:<12} {str(r['store_rating']):<6} {str(r['store_users']):<12}")
        print("=" * 100)
        print(f"Name check: match={sum(1 for x in results if x['name_check']=='match')}, partial={sum(1 for x in results if x['name_check']=='partial')}, mismatch={sum(1 for x in results if x['name_check']=='mismatch')}")


def main():
    parser = argparse.ArgumentParser(description="Verify QA scoring export against Chrome Web Store")
    parser.add_argument("--excel", type=Path, default=PROJECT_ROOT / "docs" / "qa_extensionshield" / "qa_scoring_export.xlsx", help="Path to QA export Excel")
    parser.add_argument("--limit", type=int, default=None, help="Max number of extensions to process")
    parser.add_argument("--out", type=Path, default=None, help="Output path (.md or .csv)")
    args = parser.parse_args()
    if not args.excel.exists():
        print(f"Excel not found: {args.excel}")
        sys.exit(1)
    run_verification(args.excel, args.limit, args.out)


if __name__ == "__main__":
    main()
